"""Excel Formula MCP Server — FastMCP + stdio 传输。

在 Claude Code 中提供 8 个 Excel 操作工具：
  excel_open_workbook, excel_list_sheets, excel_read_range,
  excel_get_sheet_context, excel_detect_header_row,
  excel_generate_formula, excel_explain_formula, excel_write_formula_to_cell

用法（注册到 .mcp.json）:
  {
    "mcpServers": {
      "excel-formula": {
        "command": "python3",
        "args": ["-m", "mcp_server.server"],
        "type": "stdio"
      }
    }
  }
"""

import json
import sys
from pathlib import Path

# 确保项目根在 path 中
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from mcp.server.fastmcp import FastMCP
from openpyxl import load_workbook

from engine.context_extractor import extract_sheet_context, open_and_extract
from engine.header_detector import detect_header_row
from engine.prompt_builder import build_system_prompt, build_user_message
from engine.ai_client import FormulaAIClient
from engine.response_parser import parse_response, validate_formula, FormulaResult
from engine.formula_catalog import get_function_by_name, search_functions
from mcp_server.utils.path_resolver import resolve_path
from mcp_server.utils.cell_parser import parse_cell


mcp = FastMCP("excel-formula")

# 工作簿缓存（简单内存缓存，按文件路径）
_workbook_cache: dict[str, dict] = {}

# AI 客户端（延迟初始化）
_ai_client: FormulaAIClient | None = None


def _get_ai_client() -> FormulaAIClient:
    global _ai_client
    if _ai_client is None:
        _ai_client = FormulaAIClient()
    return _ai_client


def _load_workbook(file_path: str) -> dict:
    """加载工作簿并缓存元数据。"""
    path = str(resolve_path(file_path))
    wb = load_workbook(path, data_only=False)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        sheets[name] = {
            "name": name,
            "rows": ws.max_row or 0,
            "cols": ws.max_column or 0,
        }
    info = {"file_path": str(path), "sheets": sheets}
    _workbook_cache[str(path)] = {"wb": wb, "info": info}
    return info


def _get_workbook(file_path: str):
    """获取已缓存的工作簿，未缓存则加载。"""
    path = str(resolve_path(file_path))
    if path not in _workbook_cache:
        _load_workbook(file_path)
    return _workbook_cache[path]["wb"]


# ============================================================
# MCP Tools
# ============================================================

@mcp.tool(
    name="excel_open_workbook",
    annotations={
        "readOnlyHint": True,
        "idempotentHint": True,
        "title": "打开 Excel 工作簿",
    },
)
def excel_open_workbook(file_path: str) -> str:
    """打开 .xlsx 文件并返回元数据（sheet 列表、行列数）。

    Args:
        file_path: Excel 文件的绝对路径，如 /Users/xxx/薪资调研.xlsx
    """
    try:
        info = _load_workbook(file_path)
        sheet_list = []
        for name, s in info["sheets"].items():
            sheet_list.append(f"  - {name}: {s['rows']} 行 × {s['cols']} 列")
        return (
            f"已打开: {info['file_path']}\n"
            f"共 {len(info['sheets'])} 个 Sheet:\n"
            + "\n".join(sheet_list)
        )
    except Exception as e:
        return f"打开失败: {e}"


@mcp.tool(
    name="excel_list_sheets",
    annotations={
        "readOnlyHint": True,
        "idempotentHint": True,
        "title": "列出工作表",
    },
)
def excel_list_sheets(file_path: str) -> str:
    """列出 Excel 文件中所有 sheet 的名称和行列数。

    Args:
        file_path: Excel 文件路径
    """
    try:
        info = _load_workbook(file_path)
        lines = [f"文件: {info['file_path']}"]
        for name, s in info["sheets"].items():
            lines.append(f"  {name}: {s['rows']} 行 × {s['cols']} 列")
        return "\n".join(lines)
    except Exception as e:
        return f"列出失败: {e}"


@mcp.tool(
    name="excel_read_range",
    annotations={
        "readOnlyHint": True,
        "title": "读取单元格范围",
    },
)
def excel_read_range(
    file_path: str,
    sheet: str = "",
    range_str: str = "",
    max_rows: int = 50,
) -> str:
    """读取指定工作表的数据范围。

    Args:
        file_path: Excel 文件路径
        sheet: 工作表名（留空则使用活动 sheet）
        range_str: 范围字符串如 "A1:G10"（留空则读取已使用范围，最多 max_rows 行）
        max_rows: 未指定范围时，最多读取多少行（默认50）
    """
    try:
        wb = _get_workbook(file_path)
        ws = wb[sheet] if sheet else wb.active
        rows_read = 0
        lines = [f"Sheet: {ws.title}"]

        if range_str:
            # 解析指定范围
            for row in ws[range_str]:
                values = [
                    str(c.value) if c.value is not None else ""
                    for c in row
                ]
                lines.append("\t".join(values))
        else:
            # 读取已使用范围
            for row in ws.iter_rows(
                min_row=1,
                max_row=min(ws.max_row or 1, max_rows),
                values_only=True,
            ):
                lines.append("\t".join(str(v) if v is not None else "" for v in row))
                rows_read += 1
            if ws.max_row and ws.max_row > max_rows:
                lines.append(f"...（共 {ws.max_row} 行，仅显示前 {max_rows} 行）")

        return "\n".join(lines)
    except Exception as e:
        return f"读取失败: {e}"


@mcp.tool(
    name="excel_get_sheet_context",
    annotations={
        "readOnlyHint": True,
        "idempotentHint": True,
        "title": "提取工作表上下文",
    },
)
def excel_get_sheet_context(
    file_path: str,
    sheet: str = "",
    header_row_hint: int = 0,
) -> str:
    """完整分析工作表结构：自动检测表头行、列名、数据类型、合并单元格、数据质量问题。

    这是 generate_formula 的前置步骤，建议先调用此工具了解表格结构。

    Args:
        file_path: Excel 文件路径
        sheet: 工作表名（留空则使用活动 sheet）
        header_row_hint: 手动指定表头行号（0=自动检测）
    """
    try:
        wb = _get_workbook(file_path)
        ws = wb[sheet] if sheet else wb.active
        ctx = extract_sheet_context(
            ws,
            sheet_name=ws.title,
            header_row_hint=header_row_hint if header_row_hint > 0 else None,
        )
        return ctx.describe()
    except Exception as e:
        return f"上下文提取失败: {e}"


@mcp.tool(
    name="excel_detect_header_row",
    annotations={
        "readOnlyHint": True,
        "idempotentHint": True,
        "title": "检测表头行",
    },
)
def excel_detect_header_row(
    file_path: str,
    sheet: str = "",
) -> str:
    """自动检测工作表的表头所在行号。

    Args:
        file_path: Excel 文件路径
        sheet: 工作表名（留空则使用活动 sheet）
    """
    try:
        wb = _get_workbook(file_path)
        ws = wb[sheet] if sheet else wb.active
        header_row = detect_header_row(ws)

        # 额外信息：表头行的内容
        col_headers = []
        for c in range(1, min((ws.max_column or 1) + 1, 20)):
            val = ws.cell(row=header_row, column=c).value
            if val is not None:
                col_headers.append(str(val))

        return (
            f"Sheet: {ws.title}\n"
            f"检测到的表头行: 第 {header_row} 行\n"
            f"表头内容: {' | '.join(col_headers)}"
        )
    except Exception as e:
        return f"检测失败: {e}"


@mcp.tool(
    name="excel_generate_formula",
    annotations={
        "readOnlyHint": True,
        "idempotentHint": True,
        "title": "AI 生成 Excel 公式",
    },
)
def excel_generate_formula(
    file_path: str,
    sheet: str = "",
    request: str = "",
    use_fast: bool = False,
) -> str:
    """【核心工具】将自然语言计算需求转换为 Excel 公式。

    调用 DeepSeek API，结合工作表上下文和函数目录，生成正确的 Excel 公式。
    先调用 excel_get_sheet_context 了解表结构，再调用本工具。

    Args:
        file_path: Excel 文件路径
        sheet: 工作表名（留空则使用活动 sheet）
        request: 自然语言描述的计算需求，如 "计算一线岗位薪资中位数的平均值"
        use_fast: 是否使用快速模型（简单计算可用，速度更快）
    """
    if not request:
        return "请提供计算需求描述（request 参数）"

    try:
        wb = _get_workbook(file_path)
        ws = wb[sheet] if sheet else wb.active
        ctx = extract_sheet_context(ws, ws.title)

        # 构建 prompt
        system_prompt = build_system_prompt(ctx, user_request=request)
        user_message = build_user_message(request, ctx)

        # 调用 AI
        client = _get_ai_client()
        raw = client.generate_formula_sync(
            system_prompt=system_prompt,
            user_request=user_message,
            use_fast=use_fast,
        )
        result = parse_response(raw)

        return result.describe()

    except Exception as e:
        return f"公式生成失败: {e}"


@mcp.tool(
    name="excel_explain_formula",
    annotations={
        "readOnlyHint": True,
        "idempotentHint": True,
        "title": "解释 Excel 公式",
    },
)
def excel_explain_formula(formula: str) -> str:
    """用中文解释一个 Excel 公式的含义和工作原理。

    Args:
        formula: Excel 公式字符串，如 "=VLOOKUP(A2,D5:G51,4,FALSE)"
    """
    if not formula:
        return "请提供要解释的公式"

    # 基础解析
    cleaned = formula.strip()
    if not cleaned.startswith("="):
        cleaned = "=" + cleaned

    # 提取函数名
    func_parts = cleaned.split("(")
    main_func = ""
    for part in func_parts:
        name = ""
        for c in reversed(part):
            if c.isalpha() or c.isdigit() or c == ".":
                name = c + name
            else:
                break
        if name and name[0].isalpha():
            main_func = name.upper()
            break

    # 查找函数信息
    func_entry = get_function_by_name(main_func) if main_func else None

    lines = [f"公式: {cleaned}"]

    if func_entry:
        lines.append(f"\n主要函数: {func_entry.name}（{func_entry.zh_name}）")
        lines.append(f"语法: {func_entry.syntax}")
        lines.append(f"说明: {func_entry.description_zh}")
        lines.append(f"类似示例: {func_entry.example_zh}")
        if func_entry.pitfalls:
            lines.append(f"注意事项: {'; '.join(func_entry.pitfalls)}")
    elif main_func:
        lines.append(f"\n主要函数: {main_func}（未在目录中收录，可能是不常见或自定义函数）")

    # 验证
    validation = validate_formula(cleaned)
    if validation["issues"]:
        lines.append(f"\n⚠ 潜在问题: {'; '.join(validation['issues'])}")

    return "\n".join(lines)


@mcp.tool(
    name="excel_write_formula_to_cell",
    annotations={
        "destructiveHint": True,
        "title": "写入公式到单元格",
    },
)
def excel_write_formula_to_cell(
    file_path: str,
    sheet: str = "",
    cell: str = "",
    formula: str = "",
) -> str:
    """将公式写入指定单元格并保存文件。

    ⚠ 此操作会修改文件。公式在 Excel 打开时才会计算。

    Args:
        file_path: Excel 文件路径
        sheet: 工作表名（留空则使用活动 sheet）
        cell: 目标单元格，如 "G52"
        formula: 要写入的公式，如 "=MEDIAN(G5:G51)"
    """
    if not cell or not formula:
        return "请提供目标单元格（cell）和公式（formula）"

    try:
        path = resolve_path(file_path)
        wb = _get_workbook(file_path)
        ws = wb[sheet] if sheet else wb.active

        col_letter, row_num = parse_cell(cell)

        if not formula.startswith("="):
            formula = "=" + formula

        ws[f"{col_letter}{row_num}"] = formula
        wb.save(str(path))

        return (
            f"✅ 公式已写入 {ws.title}!{col_letter}{row_num}\n"
            f"公式: {formula}\n"
            f"文件已保存: {path}\n"
            f"提示: 请在 Excel 中打开文件查看计算结果"
        )
    except Exception as e:
        return f"写入失败: {e}"


# ============================================================
# 入口
# ============================================================

def main():
    """启动 MCP Server（stdio 传输）。"""
    mcp.run(transport="stdio")


if __name__ == "__main__":
    main()
