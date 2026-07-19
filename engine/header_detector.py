"""智能表头行检测。

评分算法扫描前 N 行，综合格式和数据特征找到最可能是表头的行。
适配中文 Excel 文件常见的表头不在第1行的情况。
"""

from openpyxl.worksheet.worksheet import Worksheet
from config import MAX_HEADER_SCAN_ROWS


def detect_header_row(ws: Worksheet, max_scan: int = MAX_HEADER_SCAN_ROWS) -> int:
    """扫描工作表，返回最可能表头所在的行号（1-based）。

    评分标准：
    - 粗体: +3
    - 有背景色（非白色/非默认）: +3
    - 文本长度 2-30（典型中文表头）: +2
    - 含中文字符: +1
    - 纯数字: -2
    - 下一行有 ≥2 个数字（表头下方通常有数据）: +5
    - 该行非空单元格比例高: +0-3

    返回评分最高的行号，至少返回1。
    """
    best_row, best_score = 1, -999
    max_col = min(ws.max_column or 1, 100)
    max_row = min(ws.max_row or 1, max_scan)

    for row_idx in range(1, max_row + 1):
        score = 0
        non_empty = 0

        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            value = cell.value

            if value is None or (isinstance(value, str) and value.strip() == ""):
                continue

            non_empty += 1

            # 粗体
            if cell.font and cell.font.bold:
                score += 3

            # 有背景色
            if cell.fill and cell.fill.patternType == "solid":
                fg = cell.fill.fgColor
                if fg and fg.rgb and fg.rgb != "00000000":
                    score += 3

            # 文本特征
            if isinstance(value, str):
                clean = value.strip()
                if 2 <= len(clean) <= 30:
                    score += 2
                if any("一" <= c <= "鿿" for c in clean):
                    score += 1

            # 纯数字不是表头
            if isinstance(value, (int, float)):
                score -= 2

        # 非空比例
        if non_empty > 0 and max_col > 0:
            fill_ratio = non_empty / min(max_col, non_empty + 5)
            if fill_ratio > 0.6:
                score += 3

        # 下一行有数据（表头下方常有数值数据）
        if row_idx < ws.max_row:
            next_row_numbers = sum(
                1
                for c in range(1, max_col + 1)
                if isinstance(ws.cell(row=row_idx + 1, column=c).value, (int, float))
            )
            if next_row_numbers >= 2:
                score += 5

        if score > best_score:
            best_score = score
            best_row = row_idx

    return max(best_row, 1)


def detect_data_range(
    ws: Worksheet, header_row: int = 1
) -> tuple[int, int, int, int]:
    """根据表头行位置推断数据范围。

    返回 (first_data_row, last_data_row, first_col, last_col)。
    """
    first_data_row = header_row + 1
    last_data_row = ws.max_row or first_data_row
    first_col = 1
    last_col = ws.max_column or 1

    # 从底部向上跳过空行
    while last_data_row > first_data_row:
        has_data = any(
            ws.cell(row=last_data_row, column=c).value is not None
            for c in range(first_col, last_col + 1)
        )
        if has_data:
            break
        last_data_row -= 1

    return first_data_row, last_data_row, first_col, last_col
