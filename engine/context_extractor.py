"""工作簿上下文提取器 — 整合表头检测、合并单元格处理、类型推断。

为 prompt 构建提供结构化的表格上下文信息。
"""

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from engine.header_detector import detect_header_row, detect_data_range
from engine.merge_handler import get_merged_regions, fill_merged_values
from engine.type_detector import detect_column_type, extract_sample_values, get_data_quality_note
from config import MAX_SAMPLE_VALUES

# Excel 列号转字母
_COL_LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def _col_letter(idx: int) -> str:
    """1-based 列号 → 列字母，如 1→A, 28→AB。"""
    if idx < 1:
        return ""
    if idx <= 26:
        return _COL_LETTERS[idx - 1]
    return _col_letter((idx - 1) // 26) + _COL_LETTERS[(idx - 1) % 26]


@dataclass
class ColumnInfo:
    col_index: int  # 1-based
    col_letter: str  # A, B, C...
    name: str | None
    data_type: str
    sample_values: list
    is_merged: bool = False


@dataclass
class SheetContext:
    sheet_name: str
    header_row: int
    first_data_row: int
    last_data_row: int
    first_col: int
    last_col: int
    total_rows: int
    columns: list[ColumnInfo] = field(default_factory=list)
    merged_regions: list[dict] = field(default_factory=list)
    data_quality_notes: list[str] = field(default_factory=list)

    def describe(self) -> str:
        """生成简洁文本描述，用于 prompt 注入。"""
        lines = [
            f"Sheet: {self.sheet_name}",
            f"表头行: {self.header_row}",
            f"数据行: {self.first_data_row}-{self.last_data_row} "
            f"({self.last_data_row - self.first_data_row + 1} 行数据)",
            f"列范围: {_col_letter(self.first_col)}-{_col_letter(self.last_col)} "
            f"(共 {self.last_col - self.first_col + 1} 列)",
            "",
            "列详情:",
        ]
        for col in self.columns:
            merge_note = " [含合并单元格]" if col.is_merged else ""
            sample = (
                ", ".join(str(v)[:30] for v in col.sample_values[:3])
                if col.sample_values
                else "无数据"
            )
            lines.append(
                f"  {col.col_letter}({col.name or '无名称'}): {col.data_type}{merge_note}"
                f"  |  示例: {sample}"
            )

        if self.merged_regions:
            lines.append("")
            lines.append(f"合并区域 ({len(self.merged_regions)} 处):")
            for m in self.merged_regions[:10]:
                lines.append(f"  - {m['description']}")

        if self.data_quality_notes:
            lines.append("")
            lines.append("数据质量提示:")
            for note in self.data_quality_notes:
                lines.append(f"  ⚠ {note}")

        return "\n".join(lines)


def extract_sheet_context(
    ws: Worksheet,
    sheet_name: str = "",
    detect_header: bool = True,
    header_row_hint: int | None = None,
) -> SheetContext:
    """从工作表提取完整上下文。

    Args:
        ws: openpyxl 工作表对象
        sheet_name: 表名
        detect_header: 是否自动检测表头行
        header_row_hint: 手动指定表头行（跳过检测）
    """
    # 表头检测
    if header_row_hint is not None:
        header_row = header_row_hint
    elif detect_header:
        header_row = detect_header_row(ws)
    else:
        header_row = 1

    # 数据范围
    first_data_row, last_data_row, first_col, last_col = detect_data_range(
        ws, header_row
    )

    # 合并区域
    merged_regions = get_merged_regions(ws)

    # 逐列分析
    columns: list[ColumnInfo] = []
    data_quality_notes: list[str] = []

    for col_idx in range(first_col, last_col + 1):
        col_letter = _col_letter(col_idx)

        # 列名（从表头行读取）
        header_val = ws.cell(row=header_row, column=col_idx).value
        col_name = str(header_val).strip() if header_val is not None else None

        # 样本值（从数据行读取）
        sample_values = extract_sample_values(
            [
                ws.cell(row=r, column=col_idx).value
                for r in range(first_data_row, last_data_row + 1)
            ],
            MAX_SAMPLE_VALUES,
        )

        # 类型推断
        col_type = detect_column_type(sample_values)

        # 合并状态
        is_merged = any(
            m["min_col"] <= col_idx <= m["max_col"]
            for m in merged_regions
        )

        columns.append(
            ColumnInfo(
                col_index=col_idx,
                col_letter=col_letter,
                name=col_name,
                data_type=col_type,
                sample_values=sample_values[:5],
                is_merged=is_merged,
            )
        )

        # 数据质量提示
        note = get_data_quality_note(col_name or col_letter, col_type)
        if note:
            data_quality_notes.append(note)

    return SheetContext(
        sheet_name=sheet_name or ws.title,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        first_col=first_col,
        last_col=last_col,
        total_rows=last_data_row - first_data_row + 1,
        columns=columns,
        merged_regions=merged_regions,
        data_quality_notes=data_quality_notes,
    )


def open_and_extract(
    file_path: str | Path,
    sheet_name: str | None = None,
    header_row_hint: int | None = None,
) -> SheetContext:
    """打开 Excel 文件并提取指定 sheet 的上下文。

    如果 sheet_name 为 None，使用活动 sheet。
    """
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active
    ctx = extract_sheet_context(ws, ws.title, header_row_hint=header_row_hint)
    wb.close()
    return ctx
